
import openpyxl
import sqlite3
import uuid
from datetime import datetime, timedelta

def import_data():
    conn = sqlite3.connect('hr-platform/prisma/dev.db')
    cursor = conn.cursor()
    
    # Clear existing data
    cursor.execute("DELETE FROM Shift")
    cursor.execute("DELETE FROM KpiRecord")
    cursor.execute("DELETE FROM PromotionSale")
    cursor.execute("DELETE FROM RegistrationKpi")
    cursor.execute("DELETE FROM Employee")
    
    wb = openpyxl.load_workbook('ФЕВРАЛЬ2026.xlsx', data_only=True)
    ws = wb['График']
    
    # Get norm hours from row 1
    norm_hours = ws.cell(1, 2).value or 152
    
    # Get base year/month from row 6
    month_str = ws.cell(6, 2).value or "Февраль 2026"
    year = 2026
    month = 2  # February
    
    # Row 7 contains day numbers (1, 2, 3, ...)
    # Starting from column 4 (D)
    date_row = 7
    start_col = 4
    
    # Create dates for the month
    dates = []
    for col in range(start_col, start_col + 31):  # Max 31 days
        day_num = ws.cell(date_row, col).value
        if day_num and isinstance(day_num, (int, float)):
            try:
                date_obj = datetime(year, month, int(day_num))
                dates.append((col, date_obj))
            except:
                break
        else:
            break
    
    print(f"Found {len(dates)} days in February 2026")
    
    # Process employees starting from row 8
    employees = {}
    row_idx = 8
    
    while True:
        name_cell = ws.cell(row_idx, 2).value
        
        if not name_cell or str(name_cell).strip() == '':
            row_idx += 1
            if row_idx > 100:  # Safety limit
                break
            continue
        
        name = str(name_cell).strip()
        
        # Skip if this is a coefficient row (next row after employee)
        type_cell = ws.cell(row_idx, 3).value
        if type_cell and str(type_cell).strip().lower() == 'к':
            row_idx += 1
            continue
        
        # Determine role and salary
        role = 'ADMIN'
        base_salary = 33600
        branch = 'Дзержинского 26'
        
        # Create employee
        emp_id = str(uuid.uuid4())
        employees[name] = {
            'id': emp_id,
            'hours_row': row_idx,
            'coeff_row': row_idx + 1
        }
        
        cursor.execute(
            "INSERT INTO Employee (id, name, role, baseSalary, hourlyRate, branch, sortOrder, createdAt) VALUES (?, ?, ?, ?, 0, ?, ?, ?)",
            (emp_id, name, role, base_salary, branch, len(employees), datetime.now().isoformat())
        )
        
        # Import shifts for this employee
        for col, date_obj in dates:
            hours_val = ws.cell(row_idx, col).value
            coeff_val = ws.cell(row_idx + 1, col).value
            
            if hours_val and isinstance(hours_val, (int, float)) and hours_val > 0:
                shift_id = str(uuid.uuid4())
                shift_type = 'REGULAR'
                coefficient = float(coeff_val) if coeff_val and isinstance(coeff_val, (int, float)) else 1.0
                
                cursor.execute(
                    "INSERT INTO Shift (id, date, employeeId, type, hours, cabinetClosed, coefficient, createdAt) VALUES (?, ?, ?, ?, ?, 0, ?, ?)",
                    (shift_id, date_obj.isoformat(), emp_id, shift_type, float(hours_val), coefficient, datetime.now().isoformat())
                )
        
        # Move to next employee (skip coefficient row)
        row_idx += 2
        
        if row_idx > 100:  # Safety limit
            break
    
    # Set norm hours for February 2026
    cursor.execute(
        "INSERT OR REPLACE INTO MonthlyNorm (month, hours, createdAt) VALUES (?, ?, ?)",
        ('2026-02', norm_hours, datetime.now().isoformat())
    )
    
    conn.commit()
    
    # Print summary
    cursor.execute("SELECT COUNT(*) FROM Employee")
    emp_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM Shift")
    shift_count = cursor.fetchone()[0]
    
    print(f"✓ Imported {emp_count} employees")
    print(f"✓ Imported {shift_count} shifts")
    print(f"✓ Set norm hours: {norm_hours}")
    
    conn.close()
    wb.close()

if __name__ == "__main__":
    import_data()
