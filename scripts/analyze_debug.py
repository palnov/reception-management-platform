import pandas as pd
import openpyxl
import xlrd
import os
import sys
import msoffcrypto
import io

files = ['KPI-02.26.xls', 'ФЕВРАЛЬ2026.xlsx']

def analyze_xlsx(file_path):
    print(f"\n--- Analyzing {os.path.basename(file_path)} (via openpyxl read_only) ---")
    try:
        # data_only=True gets values not formulas, read_only=True is faster and uses less memory
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n  Sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Print first 5 rows
            print("  First 5 rows:")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                print(f"    {row}")
                row_count += 1
                if row_count >= 5:
                    break
        wb.close()
    except Exception as e:
        print(f"  Error reading {file_path}: {e}")

def analyze_xls(file_path, password=None):
    print(f"\n--- Analyzing {os.path.basename(file_path)} (via xlrd/msoffcrypto) ---")
    try:
        if password:
            print(f"  Attempting decryption with password: '{password}'")
            decrypted_workbook = io.BytesIO()
            with open(file_path, "rb") as file:
                office_file = msoffcrypto.OfficeFile(file)
                office_file.load_key(password=password)
                office_file.decrypt(decrypted_workbook)
            
            # Now read from the in-memory decrypted file
            wb = xlrd.open_workbook(file_contents=decrypted_workbook.getvalue())
        else:
            wb = xlrd.open_workbook(file_path)

        print(f"Sheet names: {wb.sheet_names()}")
        
        for sheet_name in wb.sheet_names():
            print(f"\n  Sheet: {sheet_name}")
            sheet = wb.sheet_by_name(sheet_name)
            print(f"  Rows: {sheet.nrows}, Columns: {sheet.ncols}")
            
            # Print first 5 rows
            print("  First 5 rows:")
            for i in range(min(5, sheet.nrows)):
                print(f"    {sheet.row_values(i)}")
    except Exception as e:
        print(f"  Error reading {file_path}: {e}")

def main():
    for file_name in files:
        file_path = os.path.join(os.getcwd(), file_name)
        if not os.path.exists(file_path):
            print(f"File not found: {file_name}")
            continue
            
        if file_name.endswith('.xlsx'):
            analyze_xlsx(file_path)
        elif file_name.endswith('.xls'):
            # Basic logic: if filename matches the encrypted one, use password 'q'
            password = 'q' if 'KPI' in file_name else None
            analyze_xls(file_path, password)
        else:
            print(f"Unknown format: {file_name}")

if __name__ == "__main__":
    main()
