
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('ФЕВРАЛЬ2026.xlsx', data_only=True)

# Analyze first employee sheet
first_sheet = None
for sheet_name in wb.sheetnames:
    if sheet_name not in ['Настройки', 'KPI']:
        first_sheet = sheet_name
        break

if first_sheet:
    print(f"Analyzing sheet: {first_sheet}")
    ws = wb[first_sheet]
    
    # Print first 10 rows
    for row_idx in range(1, 11):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        print(f"Row {row_idx}: {row[:10]}")  # First 10 columns

wb.close()
