
import xlrd
import msoffcrypto
import io
import openpyxl
import os

files = ['KPI-02.26.xls', 'ФЕВРАЛЬ2026.xlsx']

def decrypt_xls(file_path, password):
    decrypted_workbook = io.BytesIO()
    with open(file_path, "rb") as file:
        office_file = msoffcrypto.OfficeFile(file)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted_workbook)
    return decrypted_workbook.getvalue()

def analyze_kpi():
    try:
        print("--- KPI LOGIC ---")
        kpi_file = 'KPI-02.26.xls'
        password = 'q'
        if not os.path.exists(kpi_file):
             print(f"{kpi_file} not found")
             return

        content = decrypt_xls(kpi_file, password)
        wb = xlrd.open_workbook(file_contents=content)
        
        # Analyze 'KPI' sheet
        sheet = wb.sheet_by_name('KPI')
        print(f"Sheet: KPI ({sheet.nrows} rows)")
        headers = sheet.row_values(1)
        print(f"Headers: {headers}")

        for i in range(2, min(7, sheet.nrows)):
             row = sheet.row_values(i)
             if not str(row[0]).strip(): continue
             print(f"\nEmployee: {row[0]}")
             for h_idx, val in enumerate(row):
                 header = headers[h_idx] if h_idx < len(headers) else f"Col_{h_idx}"
                 if str(header).strip() and val not in ('', None):
                      print(f"  {header}: {val}")
        
        # Analyze 'Настройки' sheet
        print("\n--- SETTINGS ---")
        settings = wb.sheet_by_name('Настройки')
        for i in range(settings.nrows):
            row = settings.row_values(i)
            if any(row):
                print(row)
                
    except Exception as e:
        print(f"Error analyzing KPI: {e}")

def analyze_feb():
    try:
        print("\n--- FEB LOGIC ---")
        feb_file = 'ФЕВРАЛЬ2026.xlsx'
        if not os.path.exists(feb_file):
             print(f"{feb_file} not found")
             return

        wb = openpyxl.load_workbook(feb_file, read_only=True, data_only=True)
        sheet_name = 'Морозова Олеся' # Use on existing sheet
        if sheet_name in wb.sheetnames:
             print(f"Sheet: {sheet_name}")
             ws = wb[sheet_name]
             rows = list(ws.iter_rows(values_only=True))
             for i, row in enumerate(rows[:10]):
                 print(f"Row {i}: {row}")
        else:
             print(f"Sheet {sheet_name} not found. Available: {wb.sheetnames[:5]}")
        wb.close()
    except Exception as e:
        print(f"Error analyzing FEB: {e}")

if __name__ == "__main__":
    analyze_kpi()
    analyze_feb()
